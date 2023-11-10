from datetime import datetime
from pprint import pprint
import requests
import time
from openpyxl import Workbook
from tqdm import tqdm

from utils import crypto_code_extracter,is_valid_link_checker

session = requests.Session()

workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Crypto.com Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1

worksheet1.cell(row=1, column=1, value="Invalid Discord Link")
worksheet1.cell(row=1, column=2, value="Page Link")

worksheet2.cell(row=captcha_row, column=3, value="Page Link")


start = time.time()
print(start)



all_coin_data : list = []
all_coin_links : list = []


headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control": "max-age=0",
}

params = {
    'page': 0,
    'limit': '200',
}

print("Fetching all coin data...")
while True:
    params['page'] += 1
    response = session.get('https://price-api.crypto.com/price/v1/tokens', params=params, headers=headers)
    coin_data : list = response.json().get('data')
    if len(coin_data) == 0:
        break

    # extract slugs
    for coin in coin_data:
        all_coin_links.append(f"https://crypto.com/price/{coin.get('slug')}")
    break
print("Done fetching all coin data...")
print("Searching for invalid links...")

for coin_link in tqdm(all_coin_links):
    
    code = crypto_code_extracter(session=session,page_link=coin_link)
    if code is None:
        continue
    elif code == 'captcha':
        captcha_row += 1
        url_cell = worksheet2.cell(row=captcha_row, column=1, value=coin_link)
        url_cell.hyperlink = coin_link
        url_cell.style = 'Hyperlink'
    else:
        result = is_valid_link_checker(session=session,code=code)
        if not result:
            row += 1
            discord_cell = worksheet1.cell(row=row, column=1, value=f"https://discord.com/invite/{code}")
            discord_cell.hyperlink = f"https://discord.com/invite/{code}"
            discord_cell.style = 'Hyperlink'
            url_cell = worksheet1.cell(row=row, column=2, value=coin_link)
            url_cell.hyperlink = coin_link
            url_cell.style = 'Hyperlink'
    

timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

# Create the filename with the timestamp
filename = f"{timestamp}.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)