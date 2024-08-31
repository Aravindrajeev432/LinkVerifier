# read coin.json

import json
import time
import re

from tqdm import tqdm
from openpyxl import Workbook
import requests
from requests.exceptions import MissingSchema, TooManyRedirects, Timeout



with open("all_discord_links.json", "r") as f:
    coins = json.load(f)


def is_valid_link_checker(code: str) -> bool:
    """
    returns False if code is not valid
    returns True if code is valid
    """
    try:
        response = requests.get(
            f"https://discord.com/api/v10/invites/{code}",
            headers={
                "Authorization": "Bearer vh4jtqRCG5tW7NljfdihoIcBxCuspl",
            },
        )
        limit_remining: str = response.headers.get("x-ratelimit-remaining", 1)

        if int(limit_remining) <= 2:
            time.sleep(float(response.headers.get("x-ratelimit-reset-after", 20)))
        if response.status_code != 200:
            if response.status_code == 429:
                print(f"==>> limit_remining: {limit_remining}")
                print(
                    f"==>> x-ratelimit-limit: {response.headers.get('x-ratelimit-limit')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-remaining'): {response.headers.get('x-ratelimit-remaining')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-reset-after'): {response.headers.get('x-ratelimit-reset-after')}"
                )
                pass
            return False
    except (ConnectionError, Timeout, TooManyRedirects, MissingSchema) as e:
        print(e)
        return False
    return True


def check_dicord_links(discord_url: str):
    if "discordapp.com" in discord_url:
        # https://discordapp.com/invite/znjdj8q
        # 'NoneType' object has no attribute 'group'
        code_regex = r"https?:\/\/discordapp\.com\/invite\/([a-zA-Z0-9-]+)"

        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)
    if ".com" in discord_url:
        # urls ends with .com
        # extract code from url .com
        code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)

            # captcha_row += 1
            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)
    elif ".gg" in discord_url:
        # extract code from url .gg
        code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"
        except Exception as e:
            print(e)

            # captcha_row += 1

            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)


workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "CoinRanking Invalid Links"
row: int = 1
worksheet1.cell(row=1, column=1, value="Coin link")
worksheet1.cell(row=1, column=2, value="Invalid Discord Link")



for coin in tqdm(coins):
    discord_link: str = coin.get("discord_link")
    coin_name: str = coin.get("coin_slug")
    
    if discord_link is None:
        continue
    # print(discord_link)
    result: bool = check_dicord_links(discord_url=discord_link)
    if not result:
        row += 1

        slug = worksheet1.cell(row=row, column=1, value=coin_name)
        slug.hyperlink = discord_link
        slug.style = "Hyperlink"
        discord_cell = worksheet1.cell(row=row, column=2, value=discord_link)
        discord_cell.hyperlink = discord_link
        discord_cell.style = "Hyperlink"

filename = f"CoinRanking_invalid_links.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)
