

import json
import re
import time
import openpyxl
import requests
from decouple import config
from requests.exceptions import MissingSchema, TooManyRedirects, Timeout
from tqdm import tqdm

def main():
    api_key = config('APIKEY1')
    remove_old_invalid_links = input("Remove Old Invalid Links:1 for True, 0 for False ")
    
    invalid_discord_links_history_file_path = "invalid_discord_links_history.json"
    try:
        with open(invalid_discord_links_history_file_path, "r") as f:
            invalid_discord_links_history = json.load(f)
    except FileNotFoundError:
        print("File not found: invalid_discord_links_history.json")
        with open(invalid_discord_links_history_file_path, "w") as f:
            json.dump({}, f)
        with open(invalid_discord_links_history_file_path, "r") as f:
            invalid_discord_links_history = json.load(f)
    try:
        wb = openpyxl.load_workbook('opensea_invalid_links.xlsx')

        # Get the active sheet (or a specific sheet by name)
        sheet = wb.active  # or sheet = wb['Sheet1']

        # Access a specific cell's value
        # cell_value = sheet['A1'].value
        # print(f"Value in A1: {cell_value}")

        # Iterate through rows and print values
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, values_only=True):
            discord_link = row[1]
            coin_slug = row[2]
            if discord_link == "Invalid Discord Link":
                continue
            if discord_link not in invalid_discord_links_history:
                invalid_discord_links_history[discord_link] = coin_slug

        with open("invalid_discord_links_history.json", "w") as json_file:
            json.dump(invalid_discord_links_history, json_file, indent=4)

    except FileNotFoundError:
        print("File not found: coingecko_invalid_linksv4.xlsx, No problem if you don't have this file")
        pass

    

    total_limit = int(input("Enter the total limit(eg:1000,2000) min:-100 : "))
    all_collections_length = 0 
    all_discord_links_dict:dict[str:str] = {}
    count = 0
    next_page = None
    while True:
        count += 1
        print(f"Fetching Page {count}")
        time.sleep(5)
        headers = {
        'accept': 'application/json',
        'x-api-key': api_key,
        }

        params = {
            'include_hidden': 'false',
            'order_by': 'seven_day_volume',
            'limit':100
        }

        if next_page is not None:
            params['next'] = next_page

        response = requests.get('https://api.opensea.io/api/v2/collections', params=params, headers=headers)
        # print(response.headers)
        # print(response.json())
        all_collections_length += len(response.json().get('collections'))
        for collection in response.json().get('collections'):
            if collection.get('discord_url') != "":
                all_discord_links_dict[collection.get('discord_url')] = collection.get('opensea_url') 
                # print(f"==>> name: {collection.get('name')}")
                # print(f"==>> discord_url: {collection.get('discord_url')}")
                # print(f"==>> all_discord_links_dict: {all_discord_links_dict}")
        print(f"==>> all_collections_length: {len(all_discord_links_dict)}")
        if all_collections_length >= total_limit:
            break
        next_page = response.json().get('next')
        if next_page is None:
            break

    # print(f"==>> all_discord_links_dict: {all_discord_links_dict}")
    # for discord_url,opensea_url in all_discord_links_dict.items():
    #     print(f"==>> discord_url: {discord_url}")
    #     print(f"==>> opensea_url: {opensea_url}")

    
    # Checking Discord Links
    print("Checking Discord Links")
    workbook = openpyxl.Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "Opensea Invalid Links"
    
    row = 1

    worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
    worksheet1.cell(row=1, column=3, value="Page Link")

    for discord_link,opensea_url in tqdm(all_discord_links_dict.items()):
        # print(link.get('slug'))
        if discord_link is None:
            continue
        if remove_old_invalid_links == "1":
            if discord_link in invalid_discord_links_history:
                continue
        result: bool = check_dicord_links(
            discord_url=discord_link, slug=opensea_url
        )
        if not result:
            # print(link.get("slug"))
            row += 1
            
            discord_cell = worksheet1.cell(
                row=row, column=2, value=discord_link
            )
            discord_cell.hyperlink = discord_link
            discord_cell.style = "Hyperlink"
            page_link = opensea_url
            url_cell = worksheet1.cell(row=row, column=3, value=page_link)
            url_cell.hyperlink = page_link
            url_cell.style = "Hyperlink"
    if row == 1:
        print("No Invalid Links Found")
    # Create the filename with the timestamp
    filename = "opensea_invalid_links.xlsx"

    # Save the workbook to the generated filename
    workbook.save(filename)

    

def check_dicord_links(discord_url: str, slug: str):
    if ".com" in discord_url:
        # urls ends with .com
        # extract code from url .com
        code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1
            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)
    elif ".gg" in discord_url:
        # extract code from url .gg
        code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"
        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1

            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)


def is_valid_link_checker(code: str) -> bool:
    """
    returns False if code is not valid
    returns True if code is valid
    """
    try:
        response = requests.get(
            f"https://discord.com/api/v10/invites/{code}",
            headers={
                "Authorization": "Bearer vh4jtqRCG5tW7NljfdihoIcBxCuspl",
            },
        )
        limit_remining: str = response.headers.get("x-ratelimit-remaining",1)

        if int(limit_remining) <= 2:
            time.sleep(float(response.headers.get("x-ratelimit-reset-after",20)))
        if response.status_code != 200:
            if response.status_code == 429:
                print(f"==>> limit_remining: {limit_remining}")
                print(
                    f"==>> x-ratelimit-limit: {response.headers.get('x-ratelimit-limit')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-remaining'): {response.headers.get('x-ratelimit-remaining')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-reset-after'): {response.headers.get('x-ratelimit-reset-after')}"
                )
                pass
            return False
    except (ConnectionError, Timeout, TooManyRedirects, MissingSchema) as e:
        print(e)
        return False
    return True



if __name__ == '__main__':
    main()
    