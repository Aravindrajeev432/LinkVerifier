from pprint import pprint
import time
from requests import Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
from bs4 import BeautifulSoup
from tqdm import tqdm
import re
from concurrent.futures import ThreadPoolExecutor

from utils import is_valid_link
session = Session()
discord_session = Session()
from datetime import datetime
from icecream import ic
from openpyxl import Workbook
# Extract All Currency data
count: int = 1
start = 1
limit = 500
params = {
    "start": start,
    "limit": limit,
    "sortBy": "market_cap",
    "sortType": "desc",
    "convert": "USD,BTC,ETH",
    "cryptoType": "all",
    "tagType": "all",
    "audited": "false",
    "aux": "ath,atl,high24h,low24h,num_market_pairs,cmc_rank,date_added,max_supply,circulating_supply,total_supply,volume_7d,volume_30d,self_reported_circulating_supply,self_reported_market_cap",
}
all_data: list = []
crypto_data: list = []
while True:
    try:
        response = session.get(
            "https://api.coinmarketcap.com/data-api/v3/cryptocurrency/listing",
            params=params,
        )
        
    except (ConnectionError, Timeout, TooManyRedirects) as e:
        print(e)
        break

    crypto_data = response.json().get("data").get("cryptoCurrencyList")
    all_data += crypto_data
    if len(crypto_data) == 0:
        break
    if count > 0:
        params["start"] += limit

    count += 1
    time.sleep(1)
    break

headers = {
    "Authorization": "Bearer vh4jtqRCG5tW7NljfdihoIcBxCuspl",
}




# Extract All Currency slug
workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "CoinCap"
# worksheet2 = workbook.create_sheet(title="Captcha Links")
# captcha_row = 1
row = 1
worksheet1.cell(row=1, column=1, value="Currency Name")
worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")
# worksheet2.cell(row=captcha_row, column=1, value="Currency Name")
# worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
# worksheet2.cell(row=captcha_row, column=3, value="Page Link")

{'slug':["https://discord.gg/"]}
all_links : list = []
captcha_links : dict = {}
invalid_links : list = []


# for currency in tqdm(all_data):
#     pass

def process_currency(currency):
    # print(f"==>> currency: {currency.get('slug')}")
    base_url = "https://coinmarketcap.com/currencies/"
    discord_regex = r'(?:https?://)?(?:discord\.(?:[a-z]+))'
    url = f"{base_url}{currency.get('slug')}/"
    try:
        response = session.get(url)
        if response.status_code != 200:
            print(response.message)
            return
        
    except (ConnectionError, Timeout, TooManyRedirects) as e:
        print("except")
        print(e)
        return

    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    discord_links = soup.find_all("a", href=lambda href: href and "discord" in href)
    
    
    if len(discord_links) == 0:
        
        return
    # Print the links

    for url_obj in discord_links:
        discord_url = url_obj.get("href")
        if re.match(discord_regex, discord_url):
            # domain regx
            if ".com" in discord_url:
                # urls ends with .com
                # extract code from url .com
                code_regex = r'https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)'
                try:
                    code = re.search(code_regex, discord_url).group(1)
                    # result = is_valid_link(code)
                    result = False
                    if not result:
                        # print(f"==>>invalid discord_url: {discord_url}")
                        # print(f"==>>invalid code: {code}")
                        # print(f"==>> result: {result}")
                        # print("--------")
                        all_links.append({'slug':currency.get('slug'),'data':[{'discord_url': discord_url, 'code': code}]})
                        
                except Exception as e:
                    print(f"==>>145 e: {e}")
                    
                    if currency.get('slug') not in captcha_links:
                        captcha_links[currency.get('slug')] = [discord_url]
                    else:
                        captcha_links[currency.get('slug')].append(discord_url)
                
            elif ".gg" in discord_url:
                # extract code from url .gg
                code_regex = r'https?://discord\.gg/([a-zA-Z0-9-]+)'
                try:
                    code = re.search(code_regex, discord_url).group(1)
                    # result = is_valid_link(code)
                    result = False
                    if not result:
                        all_links.append({'slug':currency.get('slug'),'data':[{'discord_url': discord_url, 'code': code}]})
                except Exception as e:
                    print(f"==>>166 e: {e}")
                    if currency.get('slug') not in captcha_links:
                        captcha_links[currency.get('slug')] = [discord_url]
                    else:
                        captcha_links[currency.get('slug')].append(discord_url)
        else:
            # non discord direact urls
            response = session.get(discord_url, allow_redirects=True)
            final_url = response.url
            try:
                code_regex = r'https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)'
                code = re.search(code_regex, final_url).group(1)
                # result = is_valid_link(code)
                result = False
                if not result:
                    # row += 1
                    # worksheet1.cell(row=row, column=1, value=currency.get('name'))
                    # worksheet1.cell(row=row, column=2, value=discord_url)
                    # worksheet1.cell(row=row, column=3, value=url)
                    all_links.append({'slug':currency.get('slug'),'data':[{'discord_url': discord_url, 'code': code}]})
            except Exception as e:
                print(f"==>>192 e: {e}/{final_url}")
                # captcha_row += 1
                # worksheet2.cell(row=captcha_row, column=1, value=currency.get('name'))
                # worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                # worksheet2.cell(row=captcha_row, column=3, value=url)
                if currency.get('slug') not in captcha_links:
                        captcha_links[currency.get('slug')] = [discord_url]
                else:
                        captcha_links[currency.get('slug')].append(discord_url)
                        


def process_link(currency:dict) -> None:
    for data in currency.get('data'):
        is_valid : bool = is_valid_link(session=session,code=data.get('code'))
        if not is_valid:
            invalid_links.append({'slug':currency.get('slug'),'data':[{'discord_url': data.get('discord_url')}]})

sample =  [{'slug':'uniswap', 
            'data':[{'code': 'FCfyBSbCU5',
              'discord_url': 'https://discord.gg/FCfyBSbCU5'}]}]

print(len(all_data))
start = time.time()
print(f"==>> start: {start}")


executor = ThreadPoolExecutor(15)
elements_per_interval = 15
interval_seconds = 1
batch_count = 0

for currency in range(0, len(all_data), elements_per_interval):
    batch = all_data[currency:currency+elements_per_interval]
    [executor.submit(process_currency, c) for c in batch]
    time.sleep(interval_seconds)

executor.shutdown()

executor = ThreadPoolExecutor(15)
elements_per_interval = 14
interval_seconds = 20
batch_count = 0
ic(len(all_links))
for currency in tqdm(range(0, len(all_links), elements_per_interval)):
    batch = all_links[currency:currency+elements_per_interval]
    [executor.submit(process_link, c) for c in batch]
    
    time.sleep(interval_seconds)

executor.shutdown()
ic(len(invalid_links))

for link in invalid_links:
    for data in link.get('data'):
        row += 1
        worksheet1.cell(row=row, column=1, value=link.get('slug'))
        worksheet1.cell(row=row, column=2, value=data.get('discord_url'))
        worksheet1.cell(row=row, column=3, value=f"https://coinmarketcap.com/currencies/{link.get('slug')}/")


timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

print(f"==>> captcha_links: {captcha_links}")
# Create the filename with the timestamp
filename = f"{timestamp}.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)

total_time = time.time() - start
print(f"time taken for contacts: {total_time:.2f} seconds")
