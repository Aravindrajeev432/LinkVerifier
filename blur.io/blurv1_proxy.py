from datetime import datetime
import json
import queue
import re
import threading
import time
from openpyxl import Workbook
import requests
from utils import is_valid_link_checker

def generate_links_from_json(filename)->None:  # generating page links using the contract name
    """
    adding collection details to global queue
    """
    with open(filename, "r") as json_file:
        data = json.load(json_file)
    global q
    count = 0
    
    for item in data:
        item_name = item.get("name")
        contract_address = item.get("contractAddress")
        link = f"https://blur.io/collection/{item_name}"
        q.put({'link':link, 'contract_address':contract_address})
        
        count += 1
        if count == 5000:
            break
    

q = queue.Queue()
generate_links_from_json("blur_data.json")
workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Blur io Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1

worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")

worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")


start = time.time()
print(start)
print(q.get())



headers = {
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "Cookie": "_ga=GA1.2.683743457.1711516675; _ga_C09NSBBFNH=GS1.1.1711516674.1.1.1711516788.0.0.0; rl_page_init_referrer=RudderEncrypt%3AU2FsdGVkX18WFP6ESiYxvGknNvKZwvUyNwtCd3YLuX4%3D; rl_page_init_referring_domain=RudderEncrypt%3AU2FsdGVkX1%2BLaMjsaRNe3KRwOJU%2FNdPzZPEr7Teos0E%3D; __cf_bm=69wsiQ1wKlQ9A39205LuESy23m01MauaViCoFzlM6QY-1711606769-1.0.1.1-AWy4MFZrPvEojsrKyuT4gqYFgSfGXUdqb57IkP_FS20VoXhTXuW0rW.Mo3D3TqnxwY.qhGuWKD5N5TLMNWXRmA; rl_user_id=RudderEncrypt%3AU2FsdGVkX1%2BdgGtCGHTPdq9aNF8tPaYJoqnhE5mzBBw%3D; rl_trait=RudderEncrypt%3AU2FsdGVkX18DENqnHylpO%2FukZPqAa%2FSnp7gcoVTgdcs%3D; rl_group_id=RudderEncrypt%3AU2FsdGVkX18cH6Qq5277BnKackQcdSFvXoA98sL8tlA%3D; rl_group_trait=RudderEncrypt%3AU2FsdGVkX18Q7wxo7bV8DTX9eBGiywnku33%2FT1GnF3E%3D; rl_anonymous_id=RudderEncrypt%3AU2FsdGVkX19jZ6BcTbYFJNKsEQUAnJiZFKfvr09l%2FJlHuV3OveWxk8Y%2Fm4P6v5Xg1w7NtAPdtZGeQSWmWK4UIA%3D%3D; rl_session=RudderEncrypt%3AU2FsdGVkX1%2BtWZcyIQtof0QQMNH79M9HcO0U6%2F49q%2BMuAvtpr2Bvzl2KezVLhiNvJ73hyivdx9SQYG9PmewenkvIRTOFL3z50xi9gRacE0zqXRwozukV5piQV7qj%2BFA5KHy3FB94Y4MrobfnS6wqnA%3D%3D",
    "f-None-Match": 'W/"11e-J71MkGrfI5JPwChTpPTaXgGZh1s"',
    "Origin": "https://blur.io",
    "Referer": "https://blur.io",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
}


def check_discord_links(proxy:str):   
    global q
    global row
    global headers
    global worksheet1
    global worksheet2
    global captcha_row
    proxies = {'https': proxy}
    s = requests.session()
    s.proxies.update(proxies)
    while not q.empty():
        collection = q.get()
        link = collection['link']
        api_url = f"https://core-api.prod.blur.io/v1/collections/{collection['contract_address']}/socials"  # api url for accessing dicord link
        response = s.get(api_url, headers=headers)
        if response.status_code == 200:
            socials_data = response.json().get("socials")
            discord_links = socials_data.get("discordUrl")

        if discord_links is None:
            continue
        if len(discord_links) == 0:
            continue

        discord_url = discord_links
        if ".com" in discord_url:
            # urls ends with .com
            # extract code from url .com
            code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
            try:
                code = re.search(code_regex, discord_url).group(1)
                if code == "EhrkaCH":
                    continue
                result = is_valid_link_checker(session=s, code=code)
                if not result:
                    row += 1
                    discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet1.cell(row=row, column=3, value='link')
                    url_cell.hyperlink = link
                    url_cell.style = "Hyperlink"

            except Exception as e:
                print(e)
                captcha_row += 1

                discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"
                time.sleep(1)
        elif ".gg" in discord_url:
            print(f"==>> discord_url: {discord_url}")
            # extract code from url .gg
            code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
            try:
                code = re.search(code_regex, discord_url).group(1)
                print(f"==>> code: {code}")
                if code == "EhrkaCH":
                    continue
                result = is_valid_link_checker(session=s, code=code)
                print(f"==>> result: {result}")
                if not result:
                    row += 1
                    print(discord_url)
                    discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet1.cell(row=row, column=3, value=link)
                    url_cell.hyperlink = link
                    url_cell.style = "Hyperlink"

            except Exception as e:
                print(e)
                captcha_row += 1
                discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"
                time.sleep(1)
        else:
            # non discord direact urls
            try:
                response = s.get(discord_url, allow_redirects=True)
            except requests.exceptions.MissingSchema:

                continue
            final_url = response.url
            try:
                code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
                code = re.search(code_regex, final_url).group(1)
                if code == "EhrkaCH":
                    continue
                result = is_valid_link_checker(session=s, code=code)
                if not result:
                    row += 1
                    discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet1.cell(row=row, column=3, value=link)
                    url_cell.hyperlink = link
                    url_cell.style = "Hyperlink"
            except Exception as e:
                captcha_row += 1

                discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"
                time.sleep(1)


















file_path = "valid_ips.json"
with open(file_path, "r", encoding="utf-8") as json_file:
    proxies:list[str] = json.load(json_file)


threads = []
for proxy in proxies:
    thread = threading.Thread(target=check_discord_links, args=(proxy,))
    thread.start()
    threads.append(thread)

for thread in threads:
    thread.join()


timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

# Create the filename with the timestamp
filename = "blur_io_invalid_links.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)

total_time = time.time() - start
print(f"time taken for contacts: {total_time:.2f} seconds")
print("all done")
