import json
import requests
import time
import re
from openpyxl import Workbook
from tqdm import tqdm
from utils import is_valid_link_checker
from datetime import datetime

def fetch_all_data(initial_url, headers):
    all_data = []
    api_url = initial_url
    while api_url:
        try:
            response = requests.get(api_url, headers=headers)
            response.raise_for_status()
            data = response.json()
            items = data.get("collections", [])
            all_data.extend(items)
            if items:
                last_item = items[-1]
                contract_address = last_item.get("contractAddress")
                volume_one_day = last_item.get("volumeOneDay")
                amount = volume_one_day.get("amount") if volume_one_day else None
                if amount:
                    next_url = f"https://core-api.prod.blur.io/v1/collections/?filters=%7B%22cursor%22%3A%7B%22contractAddress%22%3A%22{contract_address}%22%2C%22volumeOneDay%22%3A%22{amount}%22%7D%2C%22sort%22%3A%22VOLUME_ONE_DAY%22%2C%22order%22%3A%22DESC%22%7D"
                    api_url = next_url
                else:
                    api_url = None
            else:
                api_url = None
            retries = 3
        except requests.exceptions.RequestException as e:
            # print("Error fetching data:", e)
            retries -= 1
            time.sleep(5)
    if retries == 0:
        print("Failed to fetch data after retries.")
    return all_data


def save_to_json(data, filename="blur_data.json"):
    with open(filename, "w") as json_file:
        json.dump(data, json_file, indent=4)


initial_url = "https://core-api.prod.blur.io/v1/collections/?filters=%7B%22sort%22%3A%22VOLUME_ONE_DAY%22%2C%22order%22%3A%22DESC%22%7D"
headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "en-US,en;q=0.9,ml;q=0.8,tr;q=0.7,ta;q=0.6",
        "Cookie": "_ga=GA1.2.683743457.1711516675; _gid=GA1.2.746631795.1711516675; _ga_C09NSBBFNH=GS1.1.1711516674.1.1.1711516788.0.0.0; rl_page_init_referrer=RudderEncrypt%3AU2FsdGVkX1%2BL2Q5Pfpu47n9LgWvZD9fEfAeoP1HHYKSq04g9RTsGwWyYN3aO94KU; rl_page_init_referring_domain=RudderEncrypt%3AU2FsdGVkX191M%2FFoyIp2gVfrh%2BYhb7X8pUfucMy1DxQ%3D; __cf_bm=QE5LK9VGiAcE3pWZViC_AvTcOyCitvbaM9nxFnByx_Q-1711519150-1.0.1.1-T8B35h0dcKjdV.HxemtpjVR_yz62taCuT.Ld3vPMoFledMk7_SO2KtUZfMPjBK5peimy2U6nK8ffM8CK1i9p8g; rl_user_id=RudderEncrypt%3AU2FsdGVkX19vYFCC%2Bp0krkU3e5HFL2h%2B5c9BXvIk8kk%3D; rl_trait=RudderEncrypt%3AU2FsdGVkX1%2FwbmmSGVj9PntOhquumJ9MHpvo8pneWes%3D; rl_group_id=RudderEncrypt%3AU2FsdGVkX18c5%2BAIuMkJyG4r8gGzLlA70A%2FGTYjGnow%3D; rl_group_trait=RudderEncrypt%3AU2FsdGVkX196I%2B8X0LtlYCb6Uu4GNJ5%2BsoHaQLf0vH0%3D; rl_anonymous_id=RudderEncrypt%3AU2FsdGVkX18coVyT%2Fy0MiW47iQ%2BLmEzKGKnyZ%2BcyUVpflvrKFj2mvZZvVZrrIuMzaZK2XY3uxh9xH5UPBALkPQ%3D%3D; rl_session=RudderEncrypt%3AU2FsdGVkX19oUlQhJ1m4%2BEFJboceNJspszLzpUfd%2FvfhAzgKpg9znPKDhBKqthmJlN%2FBnI3tZLOZG6hzitD%2FvVA936cfGTdK0LkDgddq6l6Whp1WtJXPR1kj%2FxeM6int0AfdZoXw4DuFXRZUQNDVNw%3D%3D",
        "Origin": "https://blur.io",
        "Referer": "https://blur.io/",
        "Sec-Ch-Ua": '"Chromium";v="122", "Not(A:Brand";v="24", "Google Chrome";v="122"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    }

start_time = time.time()
print("collecting collection information....")
all_data = fetch_all_data(initial_url, headers)
end_time = time.time()

if all_data:
        save_to_json(all_data)
        print("Data saved to blur_data.json")

        time_taken = end_time - start_time
        print(f"Time taken to fetch all data: {time_taken} seconds")



def generate_links_from_json(filename):  # generating page links using the contract name
    with open(filename, "r") as json_file:
        data = json.load(json_file)

    count = 0
    all_links = []
    for item in data:
        item_name = item.get("name")
        contract_address = item.get("contractAddress")
        link = f"https://blur.io/collection/{item_name}"
        all_links.append((link, contract_address))

        count += 1
        if count == 100:
            break
    return all_links


all_links = generate_links_from_json("blur_data.json")
workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Blur io Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1

worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")

worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")


start = time.time()
print(start)

headers = {
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "Cookie": "_ga=GA1.2.683743457.1711516675; _ga_C09NSBBFNH=GS1.1.1711516674.1.1.1711516788.0.0.0; rl_page_init_referrer=RudderEncrypt%3AU2FsdGVkX18WFP6ESiYxvGknNvKZwvUyNwtCd3YLuX4%3D; rl_page_init_referring_domain=RudderEncrypt%3AU2FsdGVkX1%2BLaMjsaRNe3KRwOJU%2FNdPzZPEr7Teos0E%3D; __cf_bm=69wsiQ1wKlQ9A39205LuESy23m01MauaViCoFzlM6QY-1711606769-1.0.1.1-AWy4MFZrPvEojsrKyuT4gqYFgSfGXUdqb57IkP_FS20VoXhTXuW0rW.Mo3D3TqnxwY.qhGuWKD5N5TLMNWXRmA; rl_user_id=RudderEncrypt%3AU2FsdGVkX1%2BdgGtCGHTPdq9aNF8tPaYJoqnhE5mzBBw%3D; rl_trait=RudderEncrypt%3AU2FsdGVkX18DENqnHylpO%2FukZPqAa%2FSnp7gcoVTgdcs%3D; rl_group_id=RudderEncrypt%3AU2FsdGVkX18cH6Qq5277BnKackQcdSFvXoA98sL8tlA%3D; rl_group_trait=RudderEncrypt%3AU2FsdGVkX18Q7wxo7bV8DTX9eBGiywnku33%2FT1GnF3E%3D; rl_anonymous_id=RudderEncrypt%3AU2FsdGVkX19jZ6BcTbYFJNKsEQUAnJiZFKfvr09l%2FJlHuV3OveWxk8Y%2Fm4P6v5Xg1w7NtAPdtZGeQSWmWK4UIA%3D%3D; rl_session=RudderEncrypt%3AU2FsdGVkX1%2BtWZcyIQtof0QQMNH79M9HcO0U6%2F49q%2BMuAvtpr2Bvzl2KezVLhiNvJ73hyivdx9SQYG9PmewenkvIRTOFL3z50xi9gRacE0zqXRwozukV5piQV7qj%2BFA5KHy3FB94Y4MrobfnS6wqnA%3D%3D",
    "f-None-Match": 'W/"11e-J71MkGrfI5JPwChTpPTaXgGZh1s"',
    "Origin": "https://blur.io",
    "Referer": "https://blur.io",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
}

print("starting Checking Discord Links")

session = requests.Session()
for link, contract_address in tqdm(all_links):
    api_url = f"https://core-api.prod.blur.io/v1/collections/{contract_address}/socials"  # api url for accessing dicord link
    response = requests.get(api_url, headers=headers)
    if response.status_code == 200:
        socials_data = response.json().get("socials")
        discord_links = socials_data.get("discordUrl")

    if discord_links is None:
        continue
    if len(discord_links) == 0:
        continue

    discord_url = discord_links
    if ".com" in discord_url:
        # urls ends with .com
        # extract code from url .com
        code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            if code == "EhrkaCH":
                continue
            result = is_valid_link_checker(session=session, code=code)
            if not result:
                row += 1
                discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet1.cell(row=row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)
            captcha_row += 1

            discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
            discord_cell.hyperlink = discord_url
            discord_cell.style = "Hyperlink"
            url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            url_cell.hyperlink = link
            url_cell.style = "Hyperlink"
            time.sleep(1)
    elif ".gg" in discord_url:
        # extract code from url .gg
        code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            if code == "EhrkaCH":
                continue
            result = is_valid_link_checker(session=session, code=code)
            if not result:
                row += 1
                print(discord_url)
                discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet1.cell(row=row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)
            captcha_row += 1
            discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
            discord_cell.hyperlink = discord_url
            discord_cell.style = "Hyperlink"
            url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            url_cell.hyperlink = link
            url_cell.style = "Hyperlink"
            time.sleep(1)
    else:
        # non discord direact urls
        try:
            response = session.get(discord_url, allow_redirects=True)
        except requests.exceptions.MissingSchema:

            continue
        final_url = response.url
        try:
            code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
            code = re.search(code_regex, final_url).group(1)
            if code == "EhrkaCH":
                continue
            result = is_valid_link_checker(session=session, code=code)
            if not result:
                row += 1
                discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet1.cell(row=row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"
        except Exception as e:
            captcha_row += 1

            discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
            discord_cell.hyperlink = discord_url
            discord_cell.style = "Hyperlink"
            url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            url_cell.hyperlink = link
            url_cell.style = "Hyperlink"
            time.sleep(1)

timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

# Create the filename with the timestamp
filename = "blur_io_invalid_links.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)

total_time = time.time() - start
print(f"time taken for contacts: {total_time:.2f} seconds")
print("all done")
