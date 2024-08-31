import json
import time
import re
from urllib.parse import urlparse, urlunparse
from tqdm import tqdm
from openpyxl import Workbook
from bs4 import BeautifulSoup
from icecream import ic
import requests
from requests import Session
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects,MissingSchema, InvalidSchema
from utils import is_valid_link

def main()-> None:
    discord_regex = r'(?:https?://)?(?:discord\.(?:[a-z]+))'
    session = Session()

    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "SkyNet"
    worksheet2 = workbook.create_sheet(title="Captcha Links")
    captcha_row = 1
    row = 1
    worksheet1.cell(row=1, column=1, value="Currency Name")
    worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
    worksheet1.cell(row=1, column=3, value="Page Link")
    worksheet2.cell(row=captcha_row, column=1, value="Currency Name")
    worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
    url = "https://l2beat.com/scaling/activity"
    response = requests.get(url)

    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    project_links = soup.find_all("a", href=lambda href: href and "/scaling/projects" in href)
    if len(project_links) == 0:
        return
    # Print the links
    project_links = list(set(project_links))
    print(f"==>> project_links: {len(project_links)}")
    visited:set = set({})
    # urls = [x.get("href") for x in project_links]
    # project_links = list(set(urls))
    # print(len(project_links))
    new_project_links = []
    for url_obj in project_links:
        if url_obj.get("href") in visited:
            continue
        visited.add(url_obj.get("href"))
        new_project_links.append(url_obj)
    print(f"==>> visited: {len(visited)}")
    
    
    for url_obj in tqdm(new_project_links):
        
        visited.add(url_obj.get("href"))
        time.sleep(5)
        project_url = url_obj.get("href")
        project_name = url_obj.get_text()
        ic(f"==>> project url: {project_url}")
        page_url = f"https://l2beat.com/{project_url}"
        response = requests.get(page_url)
        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        discord_links = soup.find_all("a", href=lambda href: href and "discord" in href)
        if len(discord_links) == 0:
            continue
        # Print the links
        discord_links = list(set(discord_links))
        # print(f"==>> discord_links: {discord_links}")
        visited:set = set({})
        for url_obj in discord_links:

            discord_url = url_obj.get("href")
            if discord_url not in visited:
                visited.add(discord_url)
            else:
                continue
            print(f"==>> discord_url: {discord_url}")
            
            # ic(discord_url)
            if discord_url == "https://discord.com/invite/eaVKXPmtWk" or discord_url == 'https://discord.gg/eaVKXPmtWk':
                
                continue
            parsed_url = urlparse(url)

            cleaned_url = parsed_url._replace(query='')

            final_url = urlunparse(cleaned_url)

            if re.match(discord_regex, discord_url):

                if ".com" in discord_url:
                    # urls ends with .com
                    # extract code from url .com
                    code_regex = r'https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)'
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        ic(code)
                        result = is_valid_link(session=session,code=code)
                        if not result:
                            row += 1
                            worksheet1.cell(row=row, column=1, value=project_name)
                            discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = 'Hyperlink'
                            url_cell = worksheet1.cell(row=row, column=3, value=url)
                            url_cell.hyperlink = url
                            url_cell.style = 'Hyperlink'

                    except Exception as e:
                        # ic(e)
                        captcha_row += 1
                        worksheet2.cell(row=captcha_row, column=1, value=project_name)
                        discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = 'Hyperlink'
                        url_cell = worksheet2.cell(row=captcha_row, column=3, value=url)
                        url_cell.hyperlink = url
                        url_cell.style = 'Hyperlink'
                elif ".gg" in discord_url:

                    # extract code from url .gg
                    code_regex = r'https?://discord\.gg/([a-zA-Z0-9-]+)'
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        ic(code)
                        result = is_valid_link(session=session,code=code)
                        if not result:
                            row += 1
                            worksheet1.cell(row=row, column=1, value=project_name)
                            discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = 'Hyperlink'
                            url_cell = worksheet1.cell(row=row, column=3, value=url)
                            url_cell.hyperlink = url
                            url_cell.style = 'Hyperlink'
                    except Exception as e:
                        # ic(e)
                        captcha_row += 1
                        worksheet2.cell(row=captcha_row, column=1, value=project_name)
                        discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = 'Hyperlink'
                        url_cell = worksheet2.cell(row=captcha_row, column=3, value=url)
                        url_cell.hyperlink = url
                        url_cell.style = 'Hyperlink'

            else:

                # non discord direact urls
                try:
                    response = session.get(discord_url, allow_redirects=True)
                except (MissingSchema,ConnectionError, Timeout, TooManyRedirects, InvalidSchema) as e:
                    continue
                final_url = response.url
                try:
                    code_regex = r'https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)'
                    code = re.search(code_regex, final_url).group(1)
                    ic(code)
                    result = is_valid_link(session=session,code=code)
                    if not result:
                        row += 1
                        worksheet1.cell(row=row, column=1, value=project_name)
                        discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = 'Hyperlink'
                        url_cell = worksheet1.cell(row=row, column=3, value=url)
                        url_cell.hyperlink = url
                        url_cell.style = 'Hyperlink'
                except Exception as e:
                    # ic(e)
                    captcha_row += 1
                    worksheet2.cell(row=captcha_row, column=1, value=project_name)
                    discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = 'Hyperlink'
                    url_cell = worksheet2.cell(row=captcha_row, column=3, value=url)
                    url_cell.hyperlink = url
                    url_cell.style = 'Hyperlink'
    filename = f"L2beat_invalid_links_all.xlsx"

    # Save the workbook to the generated filename
    workbook.save(filename)
    # Save the workbook to the generated filename
    workbook.save(filename)



if __name__ == "__main__":
    main()