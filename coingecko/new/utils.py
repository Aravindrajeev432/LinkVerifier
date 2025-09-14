
import json
import re
import time

import openpyxl
import requests
from requests.exceptions import MissingSchema, TooManyRedirects, Timeout

def read_contacts_from_json(json_file_path):
    try:
        with open(json_file_path, "r") as json_file:
            contacts = json.load(json_file)
        return contacts
    except FileNotFoundError:
        print(f"File not found: {json_file_path}")
        return []
    except json.JSONDecodeError:
        print(f"Error decoding JSON in file: {json_file_path}")
        return []


def check_dicord_links(discord_url: str, slug: str):
    if ".com" in discord_url:
        # urls ends with .com
        # extract code from url .com
        code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1
            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)
    elif ".gg" in discord_url:
        # extract code from url .gg
        code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"
        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1

            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)


def is_valid_link_checker(code: str) -> bool:
    """
    returns False if code is not valid
    returns True if code is valid
    """
    try:
        response = requests.get(
            f"https://discord.com/api/v10/invites/{code}",
            headers={
                "Authorization": "Bearer vh4jtqRCG5tW7NljfdihoIcBxCuspl",
            },
        )
        limit_remining: str = response.headers.get("x-ratelimit-remaining",1)

        if int(limit_remining) <= 2:
            time.sleep(float(response.headers.get("x-ratelimit-reset-after",20)))
        if response.status_code != 200:
            if response.status_code == 429:
                print(f"==>> limit_remining: {limit_remining}")
                print(
                    f"==>> x-ratelimit-limit: {response.headers.get('x-ratelimit-limit')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-remaining'): {response.headers.get('x-ratelimit-remaining')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-reset-after'): {response.headers.get('x-ratelimit-reset-after')}"
                )
                pass
            return False
    except (ConnectionError, Timeout, TooManyRedirects, MissingSchema) as e:
        print(e)
        return False
    return True



def create_invalid_discord_links_history():
    """
    python3 -c 'from utils import *; create_invalid_discord_links_history()'
    """
    with open("invalid_discord_links_history.json", "r") as json_file:
        try:
            invalid_discord_links_history:dict[str:str]= json.load(json_file)
        except FileNotFoundError:
            print("File not found: invalid_discord_links_history.json")
            invalid_discord_links_history = {}
        except json.JSONDecodeError:
            print("Error decoding JSON in file: invalid_discord_links_history.json")
            invalid_discord_links_history = {}

    
    wb = openpyxl.load_workbook('coingecko_invalid_linksv4.xlsx')

    # Get the active sheet (or a specific sheet by name)
    sheet = wb.active  # or sheet = wb['Sheet1']

    # Access a specific cell's value
    # cell_value = sheet['A1'].value
    # print(f"Value in A1: {cell_value}")

    # Iterate through rows and print values
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, values_only=True):
        page_link = row[2]
        if page_link == "Page Link":
            continue
        if page_link not in invalid_discord_links_history:
            invalid_discord_links_history[page_link] = row[1] # discord link

    with open("invalid_discord_links_history.json", "w") as json_file:
        json.dump(invalid_discord_links_history, json_file, indent=4)

