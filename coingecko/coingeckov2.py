import re
import time
from requests import Session
from icecream import ic
from tqdm import tqdm
from utils import  is_valid_link_checker
from datetime import datetime
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects
from openpyxl import Workbook
import hashlib
import json
from bs4 import BeautifulSoup


workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Coingecko Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1
worksheet1.cell(row=1, column=1, value="Currency Name")
worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")
worksheet2.cell(row=captcha_row, column=1, value="Currency Name")
worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")


session: Session = Session()
coingecko_url: str = "https://api.coingecko.com/api/v3/coins/list"
coingecko_res = session.get(coingecko_url)
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control": "max-age=0",
}


ic(len(coingecko_res.json()))
try:
    with open("coin_list.json", "r") as json_file:
        loaded_data = json.load(json_file)
    # checking urls json exist
    with open("coin_list_urls.json", "r") as json_file:
        coin_list_urls = json.load(json_file)
except FileNotFoundError:
    loaded_data = None


if loaded_data != coingecko_res.json():
    # write new file
    with open("coin_list.json", "w") as json_file:
        json.dump(coingecko_res.json(), json_file)
    coin_list_urls: list[dict] = []
    print("Building url list")
    for coin in tqdm(coingecko_res.json()):
        time.sleep(1.5)
        redirect_url = f"https://www.coingecko.com/en/search_redirect?id={coin.get('id')}&amp;type=coin"
        response = session.get(url=redirect_url, headers=headers, allow_redirects=True)
        if response.status_code != 200:
            print("rate Limit Exceeded")
            continue
        url = response.url

        coin_list_urls.append({"name": coin.get("name"), "url": url})
    else:
        with open("coin_list_urls.json", "w") as json_file:
            json.dump(coin_list_urls, json_file)
        print("Building url list Finished")
else:
    print("Checking Discord Links")
    for coin in tqdm(coin_list_urls):
        try:
            response = session.get(url=coin.get("url"), headers=headers)
            if response.status_code != 200:
                # if response.status_code == 429:
                #     time.sleep(int(response.headers.get('Retry-After')))
                print(f"==>> status_code: {response.status_code}")
                continue
        except (ConnectionError, Timeout, TooManyRedirects) as e:
            continue
        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        discord_links = soup.find_all("a", href=lambda href: href and "discord" in href)
        if len(discord_links) == 0:
            continue
        for url_obj in discord_links:
            discord_regex = r"(?:https?://)?(?:discord\.(?:[a-z]+))"
            discord_url = url_obj.get("href")
            if re.match(discord_regex, discord_url):
                if ".com" in discord_url:
                    # urls ends with .com
                    # extract code from url .com
                    code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        if code == "EhrkaCH":
                            continue
                        result = is_valid_link_checker(session=session, code=code)
                        if not result:
                            row += 1
                            worksheet1.cell(row=row, column=1, value=coin.get("name"))
                            discord_cell = worksheet1.cell(
                                row=row, column=2, value=discord_url
                            )
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = "Hyperlink"
                            url_cell = worksheet1.cell(
                                row=row, column=3, value=coin.get("url")
                            )
                            url_cell.hyperlink = coin.get("url")
                            url_cell.style = "Hyperlink"

                    except Exception as e:
                        captcha_row += 1
                        worksheet2.cell(
                            row=captcha_row, column=1, value=coin.get("name")
                        )
                        discord_cell = worksheet2.cell(
                            row=captcha_row, column=2, value=discord_url
                        )
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = "Hyperlink"
                        url_cell = worksheet2.cell(
                            row=captcha_row, column=3, value=coin.get("url")
                        )
                        url_cell.hyperlink = coin.get("url")
                        url_cell.style = "Hyperlink"
                        time.sleep(1)
                elif ".gg" in discord_url:
                    # extract code from url .gg
                    code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        if code == "EhrkaCH":
                            continue
                        result = is_valid_link_checker(session=session, code=code)
                        if not result:
                            row += 1
                            worksheet1.cell(row=row, column=1, value=coin.get("name"))
                            discord_cell = worksheet1.cell(
                                row=row, column=2, value=discord_url
                            )
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = "Hyperlink"
                            url_cell = worksheet1.cell(
                                row=row, column=3, value=coin.get("url")
                            )
                            url_cell.hyperlink = coin.get("url")
                            url_cell.style = "Hyperlink"
                    except Exception as e:
                        captcha_row += 1
                        worksheet2.cell(
                            row=captcha_row, column=1, value=coin.get("name")
                        )
                        discord_cell = worksheet2.cell(
                            row=captcha_row, column=2, value=discord_url
                        )
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = "Hyperlink"
                        url_cell = worksheet2.cell(
                            row=captcha_row, column=3, value=coin.get("url")
                        )
                        url_cell.hyperlink = coin.get("url")
                        url_cell.style = "Hyperlink"
                        time.sleep(1)
            else:
                # non discord direact urls
                response = session.get(discord_url, allow_redirects=True)
                final_url = response.url
                try:
                    code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
                    code = re.search(code_regex, final_url).group(1)
                    if code == "EhrkaCH":
                        continue
                    result = is_valid_link_checker(session=session, code=code)
                    if not result:
                        row += 1
                        worksheet1.cell(row=row, column=1, value=coin.get("name"))
                        discord_cell = worksheet1.cell(
                            row=row, column=2, value=discord_url
                        )
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = "Hyperlink"
                        url_cell = worksheet1.cell(
                            row=row, column=3, value=coin.get("url")
                        )
                        url_cell.hyperlink = coin.get("url")
                        url_cell.style = "Hyperlink"
                except Exception as e:
                    captcha_row += 1
                    worksheet2.cell(row=captcha_row, column=1, value=coin.get("name"))
                    discord_cell = worksheet2.cell(
                        row=captcha_row, column=2, value=discord_url
                    )
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet2.cell(
                        row=captcha_row, column=3, value=coin.get("url")
                    )
                    url_cell.hyperlink = coin.get("url")
                    url_cell.style = "Hyperlink"
                    time.sleep(1)

    else:
        print("Finished")
timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

# Create the filename with the timestamp
filename = f"coingecko_invalid_links_{timestamp}.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)

# json_string = json.dumps(coingecko_res.json())
# # print(json_string[:100])
# # # json_string = json.dumps(sample)
# sha256 = hashlib.sha256()
# sha256.update(json_string.encode('utf-8'))
# hash_value = sha256.hexdigest()
# print(hash_value)


# with open("coin_list_hash.txt", "w") as f:
#     f.write(str(hash_value))
