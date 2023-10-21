import time
from requests import Session
from icecream import ic
from tqdm import tqdm
from utils import coingecko_code_extracter, is_valid_link_checker
from datetime import datetime
from openpyxl import Workbook

session: Session = Session()
coingecko_url: str = "https://api.coingecko.com/api/v3/coins/list"


workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Coingecko Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1
worksheet1.cell(row=1, column=1, value="Currency Name")
worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")
worksheet2.cell(row=captcha_row, column=1, value="Currency Name")
worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")


coingecko_res = session.get(coingecko_url)
ic(len(coingecko_res.json()))
limit = 0
for coin in tqdm(coingecko_res.json()):
    limit += 1
    # if limit == 500:
    #     break
    coin_obj : dict = coingecko_code_extracter(session=session,coin_id=coin.get('id'))
    if not coin_obj:
        # not 200 when taking coingecko page
        
        continue
    if coin_obj.get('code') :
        discord_url = f"https://discord.com/invite/{coin_obj.get('code')}"
        url = coin_obj.get('url')
        # coingecko's discord link
        if coin_obj.get('code') == "EhrkaCH":
            continue
        is_valid_link : bool = is_valid_link_checker(session=session,code=coin_obj.get('code'))
        if not is_valid_link:
            
            row += 1
            worksheet1.cell(row=row, column=1, value=coin.get('name'))
            discord_cell  = worksheet1.cell(row=row, column=2, value=discord_url)
            discord_cell.hyperlink = discord_url
            discord_cell.style = 'Hyperlink'
            url_cell = worksheet1.cell(row=row, column=3, value=url)
            url_cell.hyperlink = url
            url_cell.style = 'Hyperlink'
            
    elif coin_obj.get('captcha'):
        
        captcha_row += 1
        worksheet2.cell(row=captcha_row, column=1, value=coin.get('name'))
        discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
        discord_cell.hyperlink = discord_url
        discord_cell.style = 'Hyperlink'
        url_cell = worksheet2.cell(row=captcha_row, column=3, value=url)
        url_cell.hyperlink = url
        url_cell.style = 'Hyperlink'
        time.sleep(1)
    

timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

# Create the filename with the timestamp
filename = f"coingecko_invalid_links_{timestamp}.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)