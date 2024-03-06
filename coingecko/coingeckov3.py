import time
import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from tqdm import tqdm
from utils import is_valid_link_checker
from datetime import datetime

workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Coingecko Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1

worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")

worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")


start = time.time()
print(start)

headers: dict = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Cache-Control": "max-age=0",
}
base_url = "https://www.coingecko.com"

tables: list = []
all_links: list = []
page: int = 0
valid : int = 0
invalid = 0
captcha = 0

while True:
    page += 1
    print("Processing page {0}".format(page))
    params = {"page": page}
    # if page == 2:
    #     break
    response = requests.get(base_url, headers=headers, params=params)
    if response.status_code == 429:
        page -= 1
        time.sleep(20)
        continue
    if response.status_code != 200:
        print(f"==>> response.status_code: {response.status_code}")

        print("something went wrong")
        break

    soup = BeautifulSoup(response.content, "html.parser")
    table = soup.find("div", {"class": "tw-overflow-x-auto 2lg:tw-overflow-x-visible 2lg:tw-flex 2lg:tw-justify-center"})
    if table:
        all_page_links: list = table.find_all(
            "a", href=lambda href: href and "en" in href
        )
        for link in all_page_links:
            if "www.coingecko.com" in link.get("href"):
                # redirect link
                all_links.append(link.get("href"))
            else:
                page_link = f"https://www.coingecko.com{link.get('href')}"
                all_links.append(page_link)

    else:
        break
print(f"Total coins {len(all_links)}")
if len(all_links) == 0:
    print("No coins found")
    print("Stoping program :(")
else:

    print("starting Checking Discord Links")
    # https://www.coingecko.com/en/search_redirect?id=3shares&type=coin

    session = requests.Session()
    for link in tqdm(all_links):

        response = requests.get(link, headers=headers)
        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        discord_links = soup.find_all("a", href=lambda href: href and "discord" in href)
        if len(discord_links) == 0:
            continue
        for url_obj in discord_links:
            discord_regex = r"(?:https?://)?(?:discord\.(?:[a-z]+))"
            discord_url = url_obj.get("href")
            if re.match(discord_regex, discord_url):
                if ".com" in discord_url:
                    # urls ends with .com
                    # extract code from url .com
                    code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        if code == "EhrkaCH":
                            continue
                        result = is_valid_link_checker(session=session, code=code)
                        if not result:
                            row += 1
                            print(discord_url)
                            discord_cell = worksheet1.cell(
                                row=row, column=2, value=discord_url
                            )
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = "Hyperlink"
                            url_cell = worksheet1.cell(row=row, column=3, value=link)
                            url_cell.hyperlink = link
                            url_cell.style = "Hyperlink"

                    except Exception as e:
                        print(e)
                        captcha_row += 1

                        discord_cell = worksheet2.cell(
                            row=captcha_row, column=2, value=discord_url
                        )
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = "Hyperlink"
                        url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                        url_cell.hyperlink = link
                        url_cell.style = "Hyperlink"
                        time.sleep(1)
                elif ".gg" in discord_url:
                    # extract code from url .gg
                    code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        if code == "EhrkaCH":
                            continue
                        result = is_valid_link_checker(session=session, code=code)
                        if not result:
                            row += 1
                            print(discord_url)
                            discord_cell = worksheet1.cell(
                                row=row, column=2, value=discord_url
                            )
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = "Hyperlink"
                            url_cell = worksheet1.cell(row=row, column=3, value=link)
                            url_cell.hyperlink = link
                            url_cell.style = "Hyperlink"
                    except Exception as e:
                        print(e)
                        captcha_row += 1

                        discord_cell = worksheet2.cell(
                            row=captcha_row, column=2, value=discord_url
                        )
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = "Hyperlink"
                        url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                        url_cell.hyperlink = link
                        url_cell.style = "Hyperlink"
                        time.sleep(1)
            else:
                # non discord direact urls
                try:
                    response = session.get(discord_url, allow_redirects=True)
                except requests.exceptions.MissingSchema:

                    continue
                final_url = response.url
                try:
                    code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
                    code = re.search(code_regex, final_url).group(1)
                    if code == "EhrkaCH":
                        continue
                    result = is_valid_link_checker(session=session, code=code)
                    if not result:
                        row += 1
                        print(discord_url)
                        discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = "Hyperlink"
                        url_cell = worksheet1.cell(row=row, column=3, value=link)
                        url_cell.hyperlink = link
                        url_cell.style = "Hyperlink"
                except Exception as e:
                    captcha_row += 1

                    discord_cell = worksheet2.cell(
                        row=captcha_row, column=2, value=discord_url
                    )
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                    url_cell.hyperlink = link
                    url_cell.style = "Hyperlink"
                    time.sleep(1)

    else:
        print("Finished")



    timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

    # Create the filename with the timestamp
    filename = "coingecko_invalid_links.xlsx"

    # Save the workbook to the generated filename
    workbook.save(filename)

    total_time = time.time() - start
    print(f"time taken for contacts: {total_time:.2f} seconds")
    print("all done")
