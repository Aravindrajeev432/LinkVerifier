import json
from pprint import pprint
import re
import time
import requests
from tqdm import tqdm
from requests.exceptions import MissingSchema, TooManyRedirects, Timeout
from openpyxl import Workbook
from datetime import datetime
def main():
    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "Coingecko  Invalid Links"
    
    row = 1

    worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
    worksheet1.cell(row=1, column=3, value="Page Link")


    json_file_path = "all_discord_links.json"

    # Read contacts from the JSON file
    json_data: list = read_contacts_from_json(json_file_path)

    # Print the data
    pprint(len(json_data))
    for link in tqdm(json_data):
        # print(link.get('slug'))
        result: bool = check_dicord_links(
            discord_url=link.get("discord_url"), slug=link.get("slug")
        )
        if not result:
            # print(link.get("slug"))
            row += 1
            
            discord_cell = worksheet1.cell(
                row=row, column=2, value=link.get("discord_url")
            )
            discord_cell.hyperlink = link.get("discord_url")
            discord_cell.style = "Hyperlink"
            page_link = link.get('slug')
            url_cell = worksheet1.cell(row=row, column=3, value=page_link)
            url_cell.hyperlink = page_link
            url_cell.style = "Hyperlink"

    
    

    # Create the filename with the timestamp
    filename = "coingecko_invalid_linksv4.xlsx"

    # Save the workbook to the generated filename
    workbook.save(filename)

    return


def read_contacts_from_json(json_file_path):
    try:
        with open(json_file_path, "r") as json_file:
            contacts = json.load(json_file)
        return contacts
    except FileNotFoundError:
        print(f"File not found: {json_file_path}")
        return []
    except json.JSONDecodeError:
        print(f"Error decoding JSON in file: {json_file_path}")
        return []


def check_dicord_links(discord_url: str, slug: str):
    if ".com" in discord_url:
        # urls ends with .com
        # extract code from url .com
        code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1
            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)
    elif ".gg" in discord_url:
        # extract code from url .gg
        code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"
        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1

            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)


def is_valid_link_checker(code: str) -> bool:
    """
    returns False if code is not valid
    returns True if code is valid
    """
    try:
        response = requests.get(
            f"https://discord.com/api/v10/invites/{code}",
            headers={
                "Authorization": "Bearer vh4jtqRCG5tW7NljfdihoIcBxCuspl",
            },
        )
        limit_remining: str = response.headers.get("x-ratelimit-remaining")

        if int(limit_remining) <= 2:
            time.sleep(float(response.headers.get("x-ratelimit-reset-after")))
        if response.status_code != 200:
            if response.status_code == 429:
                print(f"==>> limit_remining: {limit_remining}")
                print(
                    f"==>> x-ratelimit-limit: {response.headers.get('x-ratelimit-limit')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-remaining'): {response.headers.get('x-ratelimit-remaining')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-reset-after'): {response.headers.get('x-ratelimit-reset-after')}"
                )
                pass
            return False
    except (ConnectionError, Timeout, TooManyRedirects, MissingSchema) as e:
        print(e)
        return False
    return True


if __name__ == "__main__":
    main()
