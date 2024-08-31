from datetime import datetime
import json
import queue
import re
import threading
import time
from openpyxl import Workbook
import requests
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects, MissingSchema

start = time.time()
json_file_path = "all_discord_links.json"

def read_contacts_from_json(json_file_path):
    try:
        with open(json_file_path, "r") as json_file:
            contacts = json.load(json_file)
        return contacts
    except FileNotFoundError:
        print(f"File not found: {json_file_path}")
        return []
    except json.JSONDecodeError:
        print(f"Error decoding JSON in file: {json_file_path}")
        return []

def get_proxies()->list:
    # read valid ips from file
    file_path = "valid_ips.json"
    with open(file_path, "r", encoding="utf-8") as json_file:
        ips = json.load(json_file)
    return ips


proxies:list = get_proxies()
# print(f"==>> proxies: {proxies}")


# Read contacts from the JSON file
json_data: list = read_contacts_from_json(json_file_path)

q = queue.Queue()
# remove duplicates
slug_set:set = set()
for link in json_data:
    if link.get("slug") not in slug_set:
        slug_set.add(link.get("slug"))
        q.put(link)


workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Coingecko Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1

worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")

worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")








# print(len(discord_link_queue))


def is_valid_link_checker(session,code) -> bool:
    try:
        response = session.get(f"https://discord.com/api/v10/invites/{code}",
                               headers={"Authorization": "Bearer vh4jtqRCG5tW7NljfdihoIcBxCuspl",})
        limit_remining : str = response.headers.get('x-ratelimit-remaining',1)
        x_ratelimit_reset_after : str = response.headers.get('x-ratelimit-reset-after',20)
        print(f"==>> limit_remining: {limit_remining}")
        print(f"==>> response.headers.get('x-ratelimit-reset-after'): {x_ratelimit_reset_after}")
        if int(limit_remining) <= 2:
            
            
            time.sleep(float(x_ratelimit_reset_after))
        if response.status_code != 200:
            
            if response.status_code == 429:
                print(f"==>> limit_remining: {limit_remining}")
                print(f"==>> x-ratelimit-limit: {response.headers.get('x-ratelimit-limit')}")
                print(f"==>> response.headers.get('x-ratelimit-remaining'): {response.headers.get('x-ratelimit-remaining')}")
                print(f"==>> response.headers.get('x-ratelimit-reset-after'): {response.headers.get('x-ratelimit-reset-after')}")
                pass
            return False
    except (ConnectionError, Timeout, TooManyRedirects, MissingSchema) as e:
        print(e)
        return False
    return True





def check_discord_links(proxy:str):
    global q
    global row
    global worksheet1
    global worksheet2
    global captcha_row


    proxies = {'https': proxy}
    s = requests.session()
    s.proxies.update(proxies)
    while not q.empty():
        
        coin = q.get()
        discord_url = coin.get("discord_url")
        link =  coin.get("slug")
        if ".com" in discord_url:
            # urls ends with .com
            # extract code from url .com
            code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
            try:
                code = re.search(code_regex, discord_url).group(1)
                if code == "EhrkaCH":
                    continue
                result = is_valid_link_checker(session=s, code=code)
                if not result:
                    row += 1
                    discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet1.cell(row=row, column=3, value=link)
                    url_cell.hyperlink = link
                    url_cell.style = "Hyperlink"
                    time.sleep(1)
            except Exception as e:
                print(e)
                captcha_row += 1

                discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"
                time.sleep(1)
        elif ".gg" in discord_url:
            print(f"==>> discord_url: {discord_url}")
            # extract code from url .gg
            code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
            try:
                code = re.search(code_regex, discord_url).group(1)
                print(f"==>> code: {code}")
                if code == "EhrkaCH":
                    continue
                result = is_valid_link_checker(session=s, code=code)
                print(f"==>> result: {result}")
                if not result:
                    row += 1
                    print(discord_url)
                    discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet1.cell(row=row, column=3, value=link)
                    url_cell.hyperlink = link
                    url_cell.style = "Hyperlink"
                    time.sleep(1)
            except Exception as e:
                print(e)
                captcha_row += 1
                discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"
                time.sleep(1)
        else:
            # non discord direact urls
            try:
                response = s.get(discord_url, allow_redirects=True)
            except requests.exceptions.MissingSchema:

                continue
            final_url = response.url
            try:
                code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
                code = re.search(code_regex, final_url).group(1)
                if code == "EhrkaCH":
                    continue
                result = is_valid_link_checker(session=s, code=code)
                if not result:
                    row += 1
                    discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = "Hyperlink"
                    url_cell = worksheet1.cell(row=row, column=3, value=link)
                    url_cell.hyperlink = link
                    url_cell.style = "Hyperlink"
            except Exception as e:
                captcha_row += 1

                discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                discord_cell.hyperlink = discord_url
                discord_cell.style = "Hyperlink"
                url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
                url_cell.hyperlink = link
                url_cell.style = "Hyperlink"
                time.sleep(1)





threads = []
for proxy in proxies:
    thread = threading.Thread(target=check_discord_links, args=(proxy,))
    thread.start()
    threads.append(thread)
    

for thread in threads:
    thread.join()



timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

# Create the filename with the timestamp
filename = "coingecko_v4_invalid_links.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)

total_time = time.time() - start
print(f"time taken for contacts: {total_time:.2f} seconds")
print("all done")
