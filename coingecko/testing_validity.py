

from datetime import datetime
from openpyxl import Workbook

import requests

headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.55 Safari/537.36'
    }

# headers= {
#     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:87.0) Gecko/20100101 Firefox/87.0',
#     'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
#     'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8',
#     'Accept-Encoding':'gzip,deflate,br',
#     'Connection': 'keep-alive',
#     'Upgrade-Insecure-Requests': '1',
#     'Cache-Control': 'max-age=0'
# }

redirect_url = f"https://www.coingecko.com/en/search_redirect?id=binance-bitcoin&type=coin"

response = requests.get(url = redirect_url, headers=headers)
print(response.status_code)
if response.status_code == 200:
    print("coingecko working fine")
else:
    print("coingecko not working")
    # print(response.text)

workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "Coingecko Invalid Links"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1

worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")

worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")

timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

# Create the filename with the timestamp
filename = f"coingecko_invalid_links_{timestamp}.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)

