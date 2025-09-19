
import json
from pprint import pprint
import re
import time

import openpyxl
import requests
from requests.exceptions import MissingSchema, TooManyRedirects, Timeout
from tqdm import tqdm
from decouple import config
from openpyxl import Workbook

def main():
    pprint("Fetching Coinmarketcap Data")
    remove_old_invalid_links = input("Remove Old Invalid Links:1 for True, 0 for False ")
    
    api_key = config('API_KEY')

    with open('unique_symbols.txt', 'r') as f:
        unique_symbols = f.read().splitlines()

    params = {
    "CMC_PRO_API_KEY": api_key,
    "aux":"urls",
    'symbol': f'{unique_symbols[0]}'}  

    res = requests.get("https://pro-api.coinmarketcap.com/v2/cryptocurrency/info", params=params)
    pprint("Fetching Completed")


    # extracting discord links
    pprint("Extracting discord links And Removing Old invalid links")

    """
    create invalid_discord_links_history.json if not exist
    read coincapmarket_marketv2.xlsx and save data to invalid_discord_links_history.json
    if exist
    check the coin_data against invalid_discord_links_history.json
    remove the coin_data that exist in invalid_discord_links_history.json

    {"discord_link": "coinname"}

    """
    invalid_discord_links_history_file_path = "invalid_discord_links_history.json"
    try:
        with open(invalid_discord_links_history_file_path, "r") as f:
            invalid_discord_links_history = json.load(f)
    except FileNotFoundError:
        print("File not found: invalid_discord_links_history.json")
        with open(invalid_discord_links_history_file_path, "w") as f:
            json.dump({}, f)
        with open(invalid_discord_links_history_file_path, "r") as f:
            invalid_discord_links_history = json.load(f)
    try:
        wb = openpyxl.load_workbook('coincapmarketv2_invalid_links.xlsx')

        # Get the active sheet (or a specific sheet by name)
        sheet = wb.active  # or sheet = wb['Sheet1']

        # Access a specific cell's value
        # cell_value = sheet['A1'].value
        # print(f"Value in A1: {cell_value}")

        # Iterate through rows and print values
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, values_only=True):
            discord_link = row[1]
            coin_name = row[2]
            if discord_link == "Invalid Discord Link":
                continue
            if discord_link not in invalid_discord_links_history:
                invalid_discord_links_history[discord_link] = coin_name

        with open("invalid_discord_links_history.json", "w") as json_file:
            json.dump(invalid_discord_links_history, json_file, indent=4)

    except FileNotFoundError:
        print("File not found: coingecko_invalid_linksv4.xlsx, No problem if you don't have this file")
        pass
    # print(invalid_discord_links_history['https://discord.gg/EY9x5u74'])
    
    coin_data:dict = res.json()['data']
    data_to_save:list[dict[str,str]] = []
    for symbol_name, coin_list in coin_data.items():
        # print(coin_list)
        for coin in coin_list:
            urls_chat:list = coin.get('urls',{}).get('chat',[])
            # print(urls_chat)
            for link in urls_chat:
                if 'discord' in link:
                    # print(link)
                    if remove_old_invalid_links == "1":
                        if link in invalid_discord_links_history:
                            continue
                    data_to_save.append({"slug":coin.get('slug'),"discord_url":link})
    print(len(data_to_save))
    pprint("Extracting Completed")



    # Checking Discord Links
    pprint("Checking Discord Links")
    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "CoinCapMarket Invalid Links"
    
    row = 1

    worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
    worksheet1.cell(row=1, column=3, value="Page Link")


    

    json_data:list = data_to_save
    # Print the data
    pprint(len(json_data))
    for link in tqdm(json_data):
        # print(link.get('slug'))
        if link.get("discord_url") is None:
            continue
        result: bool = check_dicord_links(
            discord_url=link.get("discord_url"), slug=link.get("slug")
        )
        if not result:
            # print(link.get("slug"))
            row += 1
            
            discord_cell = worksheet1.cell(
                row=row, column=2, value=link.get("discord_url")
            )
            discord_cell.hyperlink = link.get("discord_url")
            discord_cell.style = "Hyperlink"
            page_link = f"https://coinmarketcap.com/currencies/{link.get('slug')}"
            url_cell = worksheet1.cell(row=row, column=3, value=page_link)
            url_cell.hyperlink = page_link
            url_cell.style = "Hyperlink"

    
    
    if row == 1:
        print("No Invalid Links Found")
    # Create the filename with the timestamp
    filename = "coincapmarketv2_invalid_links.xlsx"

    # Save the workbook to the generated filename
    workbook.save(filename)





def read_contacts_from_json(json_file_path):
    try:
        with open(json_file_path, "r") as json_file:
            contacts = json.load(json_file)
        return contacts
    except FileNotFoundError:
        print(f"File not found: {json_file_path}")
        return []
    except json.JSONDecodeError:
        print(f"Error decoding JSON in file: {json_file_path}")
        return []


def check_dicord_links(discord_url: str, slug: str):
    if ".com" in discord_url:
        # urls ends with .com
        # extract code from url .com
        code_regex = r"https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"

        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1
            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)
    elif ".gg" in discord_url:
        # extract code from url .gg
        code_regex = r"https?://discord\.gg/([a-zA-Z0-9-]+)"
        try:
            code = re.search(code_regex, discord_url).group(1)
            # print(code)
            result = is_valid_link_checker(code=code)
            return result
            # if not result:
            #     row += 1
            #     print(discord_url)
            #     discord_cell = worksheet1.cell(
            #         row=row, column=2, value=discord_url
            #     )
            #     discord_cell.hyperlink = discord_url
            #     discord_cell.style = "Hyperlink"
            #     url_cell = worksheet1.cell(row=row, column=3, value=link)
            #     url_cell.hyperlink = link
            #     url_cell.style = "Hyperlink"
        except Exception as e:
            print(e)
            print(slug)
            # captcha_row += 1

            # discord_cell = worksheet2.cell(
            #     row=captcha_row, column=2, value=discord_url
            # )
            # discord_cell.hyperlink = discord_url
            # discord_cell.style = "Hyperlink"
            # url_cell = worksheet2.cell(row=captcha_row, column=3, value=link)
            # url_cell.hyperlink = link
            # url_cell.style = "Hyperlink"
            # time.sleep(1)


def is_valid_link_checker(code: str) -> bool:
    """
    returns False if code is not valid
    returns True if code is valid
    """
    try:
        response = requests.get(
            f"https://discord.com/api/v10/invites/{code}",
            headers={
                "Authorization": "Bearer vh4jtqRCG5tW7NljfdihoIcBxCuspl",
            },
        )
        limit_remining: str = response.headers.get("x-ratelimit-remaining",1)

        if int(limit_remining) <= 2:
            time.sleep(float(response.headers.get("x-ratelimit-reset-after",20)))
        if response.status_code != 200:
            if response.status_code == 429:
                print(f"==>> limit_remining: {limit_remining}")
                print(
                    f"==>> x-ratelimit-limit: {response.headers.get('x-ratelimit-limit')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-remaining'): {response.headers.get('x-ratelimit-remaining')}"
                )
                print(
                    f"==>> response.headers.get('x-ratelimit-reset-after'): {response.headers.get('x-ratelimit-reset-after')}"
                )
                pass
            return False
    except (ConnectionError, Timeout, TooManyRedirects, MissingSchema) as e:
        print(e)
        return False
    return True


if __name__ == "__main__":
    main()