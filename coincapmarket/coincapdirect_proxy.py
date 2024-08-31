import json
import queue
import threading
import time
import requests
from requests.exceptions import ConnectionError, Timeout, TooManyRedirects,MissingSchema, InvalidSchema
from bs4 import BeautifulSoup
from tqdm import tqdm
import re
from icecream import ic
from utils import is_valid_link

s = requests.Session()
from datetime import datetime
from openpyxl import Workbook
q = queue.Queue()

batch_size: int = 500
batch_count: int = 0
ic("Please wait :)")



# Extract All Currency data
count: int = 1
start = 1
limit = 100
params = {
    "start": start,
    "limit": limit,
    "sortBy": "market_cap",
    "sortType": "desc",
    "convert": "USD,BTC,ETH",
    "cryptoType": "all",
    "tagType": "all",
    "audited": "false",
    "aux": "ath,atl,high24h,low24h,num_market_pairs,cmc_rank,date_added,max_supply,circulating_supply,total_supply,volume_7d,volume_30d,self_reported_circulating_supply,self_reported_market_cap",
}
all_data: list = []
crypto_data: list = []
while True:
    try:
        response = s.get(
            "https://api.coinmarketcap.com/data-api/v3/cryptocurrency/listing",
            params=params,
        )
    except (ConnectionError, Timeout, TooManyRedirects) as e:
        break
    crypto_data = response.json().get("data").get("cryptoCurrencyList")
    # all_data += crypto_data
    for item in crypto_data:
        q.put(item)
    ic("updating Queue")
    
    if len(crypto_data) == 0:
        break
    if count > 0:
        params["start"] += limit

    count += 1
    time.sleep(1)

# ic(len(all_data))


workbook = Workbook()
worksheet1 = workbook.active
worksheet1.title = "CoinCap"
worksheet2 = workbook.create_sheet(title="Captcha Links")
captcha_row = 1
row = 1
worksheet1.cell(row=1, column=1, value="Currency Name")
worksheet1.cell(row=1, column=2, value="Invalid Discord Link")
worksheet1.cell(row=1, column=3, value="Page Link")
worksheet2.cell(row=captcha_row, column=1, value="Currency Name")
worksheet2.cell(row=captcha_row, column=2, value="Discord Link")
worksheet2.cell(row=captcha_row, column=3, value="Page Link")
base_url = "https://coinmarketcap.com/currencies/"
discord_regex = r'(?:https?://)?(?:discord\.(?:[a-z]+))'


def coin_cheker(proxy: str):
    global q
    global row
    global discord_regex
    global base_url
    global worksheet1
    global worksheet2
    global captcha_row
    proxies = {'https': proxy}
    session = requests.session()
    s.proxies.update(proxies)
    while not q.empty():
        time.sleep(1)
        currency  = q.get()
        url = f"{base_url}{currency.get('slug')}/"
        try:
            response = session.get(url)
            
        except (ConnectionError, Timeout, TooManyRedirects) as e:
            ic("Ending program due to -->")
            ic(e)
            break

        html = response.text
        soup = BeautifulSoup(html, "html.parser")
        discord_links = soup.find_all("a", href=lambda href: href and "discord" in href)
        if len(discord_links) == 0:
            ic("skiping")
            continue
        # Print the links

        for url_obj in discord_links:

            discord_url = url_obj.get("href")
            if re.match(discord_regex, discord_url):

                # domain regx

                if ".com" in discord_url:
                    # urls ends with .com
                    # extract code from url .com
                    code_regex = r'https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)'
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        ic(f"{currency.get('slug')}--{code}")
                        result = is_valid_link(session=session,code=code)
                        if not result:
                            row += 1
                            worksheet1.cell(row=row, column=1, value=currency.get('name'))
                            discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = 'Hyperlink'
                            url_cell = worksheet1.cell(row=row, column=3, value=url)
                            url_cell.hyperlink = url
                            url_cell.style = 'Hyperlink'

                    except Exception as e:
                        # ic(e)
                        captcha_row += 1
                        worksheet2.cell(row=captcha_row, column=1, value=currency.get('name'))
                        discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = 'Hyperlink'
                        url_cell = worksheet2.cell(row=captcha_row, column=3, value=url)
                        url_cell.hyperlink = url
                        url_cell.style = 'Hyperlink'
                elif ".gg" in discord_url:

                    # extract code from url .gg
                    code_regex = r'https?://discord\.gg/([a-zA-Z0-9-]+)'
                    try:
                        code = re.search(code_regex, discord_url).group(1)
                        ic(code)
                        result = is_valid_link(session=session,code=code)
                        if not result:
                            row += 1
                            worksheet1.cell(row=row, column=1, value=currency.get('name'))
                            discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                            discord_cell.hyperlink = discord_url
                            discord_cell.style = 'Hyperlink'
                            url_cell = worksheet1.cell(row=row, column=3, value=url)
                            url_cell.hyperlink = url
                            url_cell.style = 'Hyperlink'
                    except Exception as e:
                        # ic(e)
                        captcha_row += 1
                        worksheet2.cell(row=captcha_row, column=1, value=currency.get('name'))
                        discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = 'Hyperlink'
                        url_cell = worksheet2.cell(row=captcha_row, column=3, value=url)
                        url_cell.hyperlink = url
                        url_cell.style = 'Hyperlink'

            else:

                # non discord direact urls
                try:
                    response = session.get(discord_url, allow_redirects=True)
                except (MissingSchema,ConnectionError, Timeout, TooManyRedirects, InvalidSchema) as e:
                    continue
                final_url = response.url
                try:
                    code_regex = r'https?:\/\/discord\.com\/invite\/([a-zA-Z0-9-]+)'
                    code = re.search(code_regex, final_url).group(1)
                    ic(code)
                    result = is_valid_link(session=session,code=code)
                    if not result:
                        row += 1
                        worksheet1.cell(row=row, column=1, value=currency.get('name'))
                        discord_cell = worksheet1.cell(row=row, column=2, value=discord_url)
                        discord_cell.hyperlink = discord_url
                        discord_cell.style = 'Hyperlink'
                        url_cell = worksheet1.cell(row=row, column=3, value=url)
                        url_cell.hyperlink = url
                        url_cell.style = 'Hyperlink'
                except Exception as e:
                    # ic(e)
                    captcha_row += 1
                    worksheet2.cell(row=captcha_row, column=1, value=currency.get('name'))
                    discord_cell = worksheet2.cell(row=captcha_row, column=2, value=discord_url)
                    discord_cell.hyperlink = discord_url
                    discord_cell.style = 'Hyperlink'
                    url_cell = worksheet2.cell(row=captcha_row, column=3, value=url)
                    url_cell.hyperlink = url
                    url_cell.style = 'Hyperlink'

file_path = "valid_ips.json"
with open(file_path, "r", encoding="utf-8") as json_file:
    proxies:list[str] = json.load(json_file)

try:
    threads = []
    for proxy in proxies:
        thread = threading.Thread(target=coin_cheker, args=(proxy,))
        thread.start()
        threads.append(thread)

    for thread in threads:
        thread.join()
except Exception as e:
    ic(e)
    pass

timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M")

# Create the filename with the timestamp
filename = "CoinMarketCap_invalid_links.xlsx"

# Save the workbook to the generated filename
workbook.save(filename)

total_time = time.time() - start
ic(f"time taken for contacts: {total_time:.2f} seconds")
ic("all done")
